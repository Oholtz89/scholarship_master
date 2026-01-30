"""Extract text from various document formats."""
import io
import logging
from typing import Optional

logger = logging.getLogger(__name__)


class TextExtractor:
    """Extract text content from documents."""
    
    def extract_text(self, file_id: str, file_name: str, drive_service) -> str:
        """
        Extract text from a document.
        
        Args:
            file_id: Google Drive file ID.
            file_name: Name of the file.
            drive_service: Google Drive service instance.
        
        Returns:
            Extracted text content.
        """
        file_name_lower = file_name.lower()
        
        try:
            if file_name_lower.endswith(".pdf"):
                return self._extract_pdf(file_id, drive_service)
            elif file_name_lower.endswith(".docx"):
                return self._extract_docx(file_id, drive_service)
            elif file_name_lower.endswith(".txt"):
                return self._extract_text(file_id, drive_service)
            elif file_name_lower.endswith(".xlsx") or file_name_lower.endswith(".xls"):
                return self._extract_excel(file_id, drive_service)
            else:
                logger.warning(f"Unsupported file type: {file_name}")
                return ""
        except Exception as e:
            logger.error(f"Error extracting text from {file_name}: {e}")
            raise
    
    def _extract_pdf(self, file_id: str, drive_service) -> str:
        """Extract text from PDF."""
        try:
            import PyPDF2
            
            file_content = drive_service.get_file_content(file_id)
            pdf_reader = PyPDF2.PdfReader(io.BytesIO(file_content))
            
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
            
            return text[:5000]  # Limit to first 5000 characters
        except Exception as e:
            logger.warning(f"Could not extract PDF with PyPDF2: {e}")
            return ""
    
    def _extract_docx(self, file_id: str, drive_service) -> str:
        """Extract text from DOCX."""
        try:
            from docx import Document
            
            file_content = drive_service.get_file_content(file_id)
            doc = Document(io.BytesIO(file_content))
            
            text = ""
            for para in doc.paragraphs:
                text += para.text + "\n"
            
            return text[:5000]  # Limit to first 5000 characters
        except Exception as e:
            logger.warning(f"Could not extract DOCX: {e}")
            return ""
    
    def _extract_text(self, file_id: str, drive_service) -> str:
        """Extract text from plain text file."""
        try:
            file_content = drive_service.get_file_content(file_id)
            return file_content.decode("utf-8", errors="ignore")[:5000]
        except Exception as e:
            logger.warning(f"Could not extract text: {e}")
            return ""
    
    def _extract_excel(self, file_id: str, drive_service) -> str:
        """Extract text from Excel files."""
        try:
            import openpyxl
            
            file_content = drive_service.get_file_content(file_id)
            wb = openpyxl.load_workbook(io.BytesIO(file_content))
            
            text = ""
            for sheet in wb.sheetnames[:2]:  # Process first 2 sheets
                ws = wb[sheet]
                text += f"\n=== {sheet} ===\n"
                for row in ws.iter_rows(max_rows=50, max_cols=10):
                    for cell in row:
                        if cell.value:
                            text += str(cell.value) + " | "
                    text += "\n"
            
            return text[:5000]
        except Exception as e:
            logger.warning(f"Could not extract Excel: {e}")
            return ""
